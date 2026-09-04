from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any
import re
import shutil
import sqlite3
import unicodedata

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATABASE_PATH = BASE_DIR / "fantacalcio.db"
BACKUP_DIR = BASE_DIR / "data" / "backups"

ROLE_MAP = {
    "P": "P",
    "POR": "P",
    "PORTIERE": "P",
    "PORTIERI": "P",
    "D": "D",
    "DIF": "D",
    "DIFENSORE": "D",
    "DIFENSORI": "D",
    "C": "C",
    "CC": "C",
    "CENTROCAMPISTA": "C",
    "CENTROCAMPISTI": "C",
    "A": "A",
    "ATT": "A",
    "ATTACCANTE": "A",
    "ATTACCANTI": "A",
}

FIELD_ALIASES = {
    "player_id": [
        "playerid",
        "id",
        "idgiocatore",
        "codice",
        "codicegiocatore",
    ],
    "nome": [
        "nome",
        "giocatore",
        "player",
        "calciatore",
    ],
    "squadra": [
        "squadra",
        "team",
        "club",
    ],
    "ruolo": [
        "ruolo",
        "role",
        "r",
    ],
    "ruolo_originale": [
        "ruolomantra",
        "ruolooriginale",
        "mantra",
    ],
    "presenze": [
        "presenze",
        "pres",
        "pg",
        "partitegiocate",
    ],
    "media_voto": [
        "mediavoto",
        "mv",
        "media",
    ],
    "fantamedia": [
        "fantamedia",
        "fm",
    ],
    "gol_fatti": [
        "golfatti",
        "gol",
        "gf",
    ],
    "gol_subiti": [
        "golsubiti",
        "gs",
    ],
    "rigori_segnati": [
        "rigorisegnati",
        "rsegnati",
        "rs",
    ],
    "rigori_sbagliati": [
        "rigorisbagliati",
        "rsbagliati",
        "rr",
    ],
    "assist": [
        "assist",
        "ass",
    ],
    "ammonizioni": [
        "ammonizioni",
        "ammonizione",
        "amm",
    ],
    "espulsioni": [
        "espulsioni",
        "espulsione",
        "esp",
    ],
    "autogol": [
        "autogol",
        "ag",
    ],
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("’", "'").replace("‘", "'")
    return re.sub(r"s+", " ", text)


def normalized_key(value: Any) -> str:
    text = clean_text(value).lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(
        character
        for character in text
        if unicodedata.category(character) != "Mn"
    )
    return re.sub(r"[^a-z0-9]", "", text)


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return default


def normalize_role(value: Any) -> str:
    raw = clean_text(value).upper()
    if not raw:
        return ""

    first_role = re.split(r"[/,;s]+", raw)[0]
    return ROLE_MAP.get(first_role, ROLE_MAP.get(raw, ""))


def find_column_indexes(headers: list[Any]) -> dict[str, int]:
    normalized_headers = {
        normalized_key(header): index
        for index, header in enumerate(headers)
        if clean_text(header)
    }

    indexes: dict[str, int] = {}

    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized_headers:
                indexes[field] = normalized_headers[alias]
                break

    return indexes


def get_cell(row: tuple[Any, ...], indexes: dict[str, int], field: str) -> Any:
    index = indexes.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def find_best_sheet(workbook) -> tuple[Any, int, dict[str, int]]:
    best_sheet = None
    best_header_row = 0
    best_indexes: dict[str, int] = {}
    best_score = -1

    for worksheet in workbook.worksheets:
        max_header_row = min(20, worksheet.max_row)

        for header_row in range(1, max_header_row + 1):
            headers = [
                worksheet.cell(header_row, column).value
                for column in range(1, worksheet.max_column + 1)
            ]
            indexes = find_column_indexes(headers)

            score = 0
            for required_field in ("nome", "squadra", "ruolo"):
                if required_field in indexes:
                    score += 10

            score += len(indexes)

            if score > best_score:
                best_sheet = worksheet
                best_header_row = header_row
                best_indexes = indexes
                best_score = score

    if best_sheet is None:
        raise ValueError("Nessun foglio leggibile trovato nel file Excel.")

    missing = [
        field
        for field in ("nome", "squadra", "ruolo")
        if field not in best_indexes
    ]

    if missing:
        raise ValueError(
            "Colonne obbligatorie non riconosciute: "
            + ", ".join(missing)
            + ". Servono almeno nome, squadra e ruolo."
        )

    return best_sheet, best_header_row, best_indexes


def read_excel_players(file_path: str | Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    worksheet, header_row, indexes = find_best_sheet(workbook)

    players: list[dict[str, Any]] = []
    skipped_rows = 0
    missing_role_rows = 0
    used_ids: set[int] = set()
    generated_id = 900000

    for values in worksheet.iter_rows(
        min_row=header_row + 1,
        values_only=True,
    ):
        name = clean_text(get_cell(values, indexes, "nome"))
        team = clean_text(get_cell(values, indexes, "squadra"))
        role_raw = get_cell(values, indexes, "ruolo")
        role = normalize_role(role_raw)

        if not name and not team and not role_raw:
            continue

        if not name or not team:
            skipped_rows += 1
            continue

        if not role:
            missing_role_rows += 1
            skipped_rows += 1
            continue

        player_id = to_int(get_cell(values, indexes, "player_id"), 0)

        if player_id <= 0 or player_id in used_ids:
            generated_id += 1
            player_id = generated_id

        used_ids.add(player_id)

        players.append(
            {
                "player_id": player_id,
                "nome": name,
                "squadra": team,
                "ruolo": role,
                "ruolo_originale": clean_text(
                    get_cell(values, indexes, "ruolo_originale")
                )
                or clean_text(role_raw),
                "presenze": to_int(get_cell(values, indexes, "presenze")),
                "media_voto": to_float(get_cell(values, indexes, "media_voto")),
                "fantamedia": to_float(get_cell(values, indexes, "fantamedia")),
                "gol_fatti": to_int(get_cell(values, indexes, "gol_fatti")),
                "gol_subiti": to_int(get_cell(values, indexes, "gol_subiti")),
                "rigori_segnati": to_int(
                    get_cell(values, indexes, "rigori_segnati")
                ),
                "rigori_sbagliati": to_int(
                    get_cell(values, indexes, "rigori_sbagliati")
                ),
                "assist": to_int(get_cell(values, indexes, "assist")),
                "ammonizioni": to_int(
                    get_cell(values, indexes, "ammonizioni")
                ),
                "espulsioni": to_int(get_cell(values, indexes, "espulsioni")),
                "autogol": to_int(get_cell(values, indexes, "autogol")),
                "disponibile_serie_a": 1,
                "stato_dato": "import_excel",
                "file_origine": Path(file_path).name,
            }
        )

    role_counts = {"P": 0, "D": 0, "C": 0, "A": 0}
    for player in players:
        role_counts[player["ruolo"]] += 1

    summary = {
        "sheet_name": worksheet.title,
        "header_row": header_row,
        "recognized_fields": sorted(indexes.keys()),
        "valid_players": len(players),
        "skipped_rows": skipped_rows,
        "missing_role_rows": missing_role_rows,
        "role_counts": role_counts,
    }

    return players, summary


def backup_database() -> Path:
    if not DATABASE_PATH.exists():
        raise FileNotFoundError(f"Database non trovato: {DATABASE_PATH}")

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = BACKUP_DIR / f"fantacalcio_pre_import_{timestamp}.db"

    shutil.copy2(DATABASE_PATH, destination)
    return destination


def replace_current_players(players: list[dict[str, Any]]) -> int:
    if not players:
        raise ValueError("Nessun giocatore valido da importare.")

    connection = sqlite3.connect(DATABASE_PATH)

    try:
        cursor = connection.cursor()
        cursor.execute("BEGIN")
        cursor.execute("DELETE FROM current_players")

        cursor.executemany(
            """
            INSERT INTO current_players (
                player_id,
                nome,
                squadra,
                ruolo,
                ruolo_originale,
                presenze,
                media_voto,
                fantamedia,
                gol_fatti,
                gol_subiti,
                rigori_segnati,
                rigori_sbagliati,
                assist,
                ammonizioni,
                espulsioni,
                autogol,
                disponibile_serie_a,
                stato_dato,
                file_origine
            )
            VALUES (
                :player_id,
                :nome,
                :squadra,
                :ruolo,
                :ruolo_originale,
                :presenze,
                :media_voto,
                :fantamedia,
                :gol_fatti,
                :gol_subiti,
                :rigori_segnati,
                :rigori_sbagliati,
                :assist,
                :ammonizioni,
                :espulsioni,
                :autogol,
                :disponibile_serie_a,
                :stato_dato,
                :file_origine
            )
            """,
            players,
        )

        connection.commit()
        return len(players)

    except Exception:
        connection.rollback()
        raise

    finally:
        connection.close()
